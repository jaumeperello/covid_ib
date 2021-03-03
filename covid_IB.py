#!/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
import datetime
import numpy as np
from src.arcgis_scraper import arcgis_has_changes
from src.gitcovid19_ib import hospital_has_changes
from src.goib_xlsx_dowload import goib_xlsx_has_changes

import logging

logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)


def get_filename_extension(path):
    *other, filename = path.split('/')
    *other, extension = filename.split('.')
    return filename, extension


def merge_files_to_csv(inputpath="dowload/", outputpath="data/"):
    # getting cases and tp7d from IBgov xlsx
    wb = load_workbook(filename=f"{inputpath}gov_xlsx/goib_covid.xlsx",
                       data_only=True, read_only=True)
    # get cases from xlsx
    sheet = wb['PDIA+xMUNICIPI']
    row_header = 0
    for row in sheet.iter_rows():
        if row[1].value == 'MUNICIPI_NOU':
            row_header = row[1].row
            break

    row_last_mun = 0
    for row in sheet.iter_rows(min_row=row_header):
        if row[1].value and row[1].value == 'TOTAL ILLES BALEARS':
            row_last_mun = row[1].row
            break

    scope = ['mallorca', 'menorca', 'eivissa', 'formentera']
    municipis_df = {}

    # getting deaceased, active and recovered from Arcgis
    for illa in scope:
        municipis_df[illa] = pd.read_csv(f"{inputpath}/arcgis/{illa}_total.csv")
        municipis_df[illa].set_index(['date', 'region_code'], inplace=True)

    # getting hospitalized, intensivecare from https://github.com/druizaguilera/covid19_IB

    ib_df = pd.read_csv(f"{inputpath}/covid19_IB/covid19_IB.csv")
    ib_df.set_index(['date', 'region_code'], inplace=True)

    yesterday = datetime.date.today()
    yesterday = datetime.datetime(yesterday.year, yesterday.month, yesterday.day)
    yesterday = yesterday - datetime.timedelta(days=1)

    filename = f"{outputpath}renameRegions.csv"
    renameRegions = pd.read_csv(filename, index_col=0, header=None, squeeze=True).to_dict()

    filename = f"{outputpath}regionsCodes.csv"
    regionsCodes = pd.read_csv(filename, index_col=0, header=None, squeeze=True).to_dict()

    filename = f"{outputpath}islandCodes.csv"
    islandCodes = pd.read_csv(filename, index_col=0, header=None, squeeze=True).to_dict()

    # Getting cases data by municipis into a dicc

    # find last row
    col_yesterday = 300
    while True:
        difference = yesterday - sheet.cell(row=row_header, column=col_yesterday).value
        if difference.days < 1 or not sheet.cell(row=row_header, column=col_yesterday + 1).value:
            break
        col_yesterday = col_yesterday + 1
    dates = sheet[row_header]
    data = {}
    for illa in scope:
        data[illa] = []
    dataBalears = []
    dades = {}
    for i in range(3, col_yesterday):
        dades[dates[i].value.strftime("%Y-%m-%d")] = {}
    for row in sheet.iter_rows(min_row=row_header + 1, max_row=row_last_mun, min_col=0, max_col=col_yesterday, values_only=True):
        if row[0] == 'S/D' or row[1] == 'ILLES' or row[1] == 'Desconeguda':
            continue
        if row[0] is None and row[1] is None:
            continue
        total_casos = 0
        for i in range(3, col_yesterday):
            illa = None
            nom = None
            name = row[1]
            total_casos = row[i] + total_casos
            if row[1].lower() in scope and row[0] is None:
                illa = 'Balears'
                nom = row[1]
            elif row[1] == 'TOTAL ILLES BALEARS':
                illa = 'Balears'
                nom = 'total-balears'
                name = 'Balears'
            elif row[1] != 'S/D':
                illa = row[0]
                nom = renameRegions[row[1]]
            dades[dates[i].value.strftime("%Y-%m-%d")][name] = {
                'nom': nom,
                'illa': illa,
                'casos': total_casos
            }

    # Reading TP7ID #

    sheet = wb['TP7D']
    row_header = 0

    # islands rows
    for row in sheet.iter_rows():
        if row[0].value == 'Tp 7D PER ILLES':
            row_header = row[0].row
            break
    row_first_isl = row_header + 2
    row_last_isl = row_first_isl + 5

    # municipis rows
    for row in sheet.iter_rows():
        if row[0].value == 'Tp 7D PER MUNICIPIS':
            row_header = row[1].row
            break

    row_first_mun = row_header + 1

    row_last_mun = 0
    for row in sheet.iter_rows(min_row=row_header):
        if row[1].value is None or row[0].value == 'Tp 7D PER EAP':
            break
        row_last_mun = row[0].row

    col_yesterday = 300
    while True:
        difference = yesterday - sheet.cell(row=row_header, column=col_yesterday).value
        if difference.days < 1 or not sheet.cell(row=row_header, column=col_yesterday + 1).value:
            break
        col_yesterday = col_yesterday + 1
    dates = sheet[row_header]
    tp_illes = {}
    tp_mun = {}
    for i in range(1, col_yesterday):
        tp_illes[dates[i].value.strftime("%Y-%m-%d")] = {}
        tp_mun[dates[i].value.strftime("%Y-%m-%d")] = {}

    # Get Islands data
    for row in sheet.iter_rows(min_row=row_first_isl, max_row=row_last_isl, min_col=0, max_col=col_yesterday, values_only=True):
        for i in range(1, col_yesterday):
            if row[0] != 'Desconeguda':
                illa = row[0]
                if illa == 'TOTAL ILLES BALEARS':
                    illa = 'total-balears'
                tp_illes[dates[i].value.strftime("%Y-%m-%d")][illa] = row[i]
    # Get Municipis data
    for row in sheet.iter_rows(min_row=row_first_mun, max_row=row_last_mun, min_col=0, max_col=col_yesterday, values_only=True):
        if row[0] != 'Desconeguda':
            for i in range(1, col_yesterday):
                tp_mun[dates[i].value.strftime("%Y-%m-%d")][renameRegions[row[0]]] = row[i]

    for illa in scope:
        data[illa].append(['date', 'region_code', 'region', 'illa', 'cases', 'recovered', 'active_cases', 'deceased', 'tp7d'])
    dataBalears.append(['date', 'region_code', 'region', 'cases', 'recovered', 'active_cases', 'deceased', 'tp7d', 'hospitalized', 'intensivecare'])
    previous_dates = []
    for date in dades.keys():
        logging.info(f"processing: {date}")
        illes = {
            'Mallorca': {},
            'Menorca': {},
            'Eivissa': {},
            'Formentera': {},
            'total-balears': {}
        }

        for illa in illes.keys():
            illes[illa] = {
                'cases': 0,
                'recovered': 0,
                'active_cases': 0,
                'deceased': 0,
            }

        for region in dades[date].keys():
            recovered, active, deceased = (0, 0, 0)
            dada = dades[date][region]
            illa = dada['illa'].lower()
            tp7d = 0
            regionCode = 0
            if illa != "balears":
                regionCode = int(regionsCodes[dada['nom']])
            if date in tp_mun.keys() and illa != 'balears':
                tp7d = tp_mun[date][dada['nom']]

            # adding recovered active and deceased from arcgis starting from 2020-10-6
            if datetime.datetime.strptime(date, '%Y-%m-%d') > datetime.datetime.strptime("2020-10-6", '%Y-%m-%d'):
                if illa != "balears" and municipis_df[illa].index.isin([(date, regionCode)]).any():
                    recovered = max(0, dada['casos'] - municipis_df[illa].loc[(date, regionCode)]['active_cases'])
                    active = municipis_df[illa].loc[(date, regionCode)]['active_cases']
                    deceased = municipis_df[illa].loc[(date, regionCode)]['deceased']
                elif len(previous_dates) > 0 and illa != "balears":
                    for previous_date in previous_dates[::-1]:
                        if municipis_df[illa].index.isin([(previous_date, regionCode)]).any():
                            recovered = max(0, dada['casos'] - municipis_df[illa].loc[(previous_date, regionCode)]['active_cases'])
                            active = municipis_df[illa].loc[(previous_date, regionCode)]['active_cases']
                            deceased = municipis_df[illa].loc[(previous_date, regionCode)]['deceased']
                            break

            # total data for IB, and every island
            if illa != 'balears':
                illes[dada['illa']]['recovered'] += recovered
                illes[dada['illa']]['active_cases'] += active
                illes[dada['illa']]['deceased'] += deceased
                illes['total-balears']['recovered'] += recovered
                illes['total-balears']['active_cases'] += active
                illes['total-balears']['deceased'] += deceased
            elif region == 'Balears':
                illes['total-balears']['cases'] = dada['casos']
                illes['total-balears']['hospitalized'] = float("NaN")
                illes['total-balears']['intensivecare'] = float("NaN")
                if ib_df.index.isin([(date, 0)]).any():
                    illes['total-balears']['hospitalized'] = int(ib_df.loc[(date, 0)]['active_hospital_admissions']) + int(ib_df.loc[(date, 0)]['active_icu'])
                    illes['total-balears']['intensivecare'] = int(ib_df.loc[(date, 0)]['active_icu'])
            else:
                illes[region]['cases'] = dada['casos']
                illes[region]['hospitalized'] = float("NaN")
                illes[region]['intensivecare'] = float("NaN")
                # adding hospitalized and intensivecare if regions is IB or an island
                if ib_df.index.isin([(date, islandCodes[region])]).any() and not np.isnan(ib_df.loc[(date, islandCodes[region])]['active_hospital_admissions']):
                    illes[region]['hospitalized'] = int(ib_df.loc[(date, islandCodes[region])]['active_hospital_admissions']) + int(ib_df.loc[(date, islandCodes[region])]['active_icu'])
                    illes[region]['intensivecare'] = int(ib_df.loc[(date, islandCodes[region])]['active_icu'])

            if illa != 'balears':
                data[illa.lower()].append([date, regionCode, dada['nom'], dada['illa'], dada['casos'], recovered, active, deceased, tp7d])

        for illa in illes.keys():
            tp7d = 0
            if date in tp_illes.keys():
                tp7d = tp_illes[date][illa]
            if illa != 'total-balears':
                data[illa.lower()].append([date, 0, "total-" + illa.lower(), illa, illes[illa]['cases'], illes[illa]['recovered'], illes[illa]['active_cases'], illes[illa]['deceased'], tp7d])
            dataBalears.append([date, islandCodes[illa], illa, illes[illa]['cases'], illes[illa]['recovered'], illes[illa]['active_cases'], illes[illa]['deceased'], tp7d, illes[illa]['hospitalized'], illes[illa]['intensivecare']])
        previous_dates.append(date)

    for illa in scope:
        np.savetxt(f"{outputpath}{illa}_total.csv", data[illa], delimiter=",", fmt='%s')
        logging.info(f"File converted. Saved at {outputpath}{illa}_total.csv")
    np.savetxt(f"{outputpath}balears_total.csv", dataBalears, delimiter=",", fmt='%s')
    logging.info(f"File converted. Saved at {outputpath}balears_total.csv")
    logging.info('done')


def get_csv(outputpath="data/", inputpath="download/", force=False):

    arcgis = arcgis_has_changes(f"{inputpath}arcgis/", "arcgis_dades/", "arcgis_cvs/")
    hospital = hospital_has_changes(f"{inputpath}covid19_IB/")
    goib = goib_xlsx_has_changes(f"{inputpath}gov_xlsx/")
    if arcgis or hospital or goib or force:
        merge_files_to_csv(inputpath, outputpath)
        logging.info("New files generated")
        return True
    logging.info("No New files")
    return False


if __name__ == "__main__":
    get_csv("data/", "download/", True)
